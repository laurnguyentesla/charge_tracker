"""Sanitize Microsoft Forms exports for GitHub Pages.

Anyone in an export is treated as having submitted. Emails, IDs, start
times, written answers, links, and file-upload URLs are stripped before
publishing. Duplicate names keep the latest completion time.
"""

from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime, timezone
from io import StringIO
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
WORKSPACE = REPO.parent
DOWNLOADS = Path.home() / "Downloads"
SOURCE_DIR = REPO / "source"

SKIP_NAMES = {"step tracker"}
SKIP_EXT = {".m4a", ".mp4", ".jpg", ".jpeg", ".png", ".mov"}

DATASETS = [
    {
        "key": "spotlight",
        "title": "Employee Spotlight",
        "subtitle": "Newsletter questionnaire",
        "source_label": "Employee Spotlight Questionnaire",
        "output": REPO / "data" / "charges.json",
        "copy_name": "employee-spotlight",
        "match": lambda name: "spotlight" in name,
        "extras": False,
    },
    {
        "key": "guild",
        "title": "Grow with Guild",
        "subtitle": "Guild program stories",
        "source_label": "Grow with Guild",
        "output": REPO / "data" / "guild.json",
        "copy_name": "grow-with-guild",
        "match": lambda name: "guild" in name and "spotlight" not in name,
        "extras": True,
    },
    {
        "key": "pathways",
        "title": "Pathways to Success",
        "subtitle": "Newsletter questionnaire",
        "source_label": "Pathways to Success - Newsletter",
        "output": REPO / "data" / "pathways.json",
        "copy_name": "pathways-to-success",
        "match": lambda name: "pathways" in name,
        "extras": False,
    },
]


def parse_datetime(value: object) -> datetime:
    if value is None or value == "":
        return datetime.min
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    raw = str(value).strip()
    for fmt in (
        "%m/%d/%y %H:%M:%S",
        "%m/%d/%y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return datetime.min


def isoformat(value: datetime | None) -> str | None:
    if value is None or value is datetime.min:
        return None
    return value.strftime("%Y-%m-%dT%H:%M:%S")


def display_time(value: datetime | None) -> str:
    if value is None or value is datetime.min:
        return ""
    hour = value.strftime("%I").lstrip("0") or "0"
    return f"{value.strftime('%b')} {value.day}, {value.year} {hour}:{value.strftime('%M %p')}"


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y %H:%M:%S")
    return str(value).replace("\xa0", " ").strip()


def is_usable(path: Path) -> bool:
    if not path.is_file() or path.suffix.casefold() in SKIP_EXT:
        return False
    if path.suffix.casefold() not in {".csv", ".xlsx", ".xlsm"}:
        return False
    lowered = path.name.casefold()
    return not any(skip in lowered for skip in SKIP_NAMES)


def find_source(dataset: dict) -> Path | None:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    match = dataset["match"]
    downloads = [
        path
        for path in DOWNLOADS.iterdir()
        if is_usable(path) and match(path.name.casefold())
    ]
    if downloads:
        csvs = [path for path in downloads if path.suffix.lower() == ".csv"]
        newest = max(csvs or downloads, key=lambda path: path.stat().st_mtime)
        dest = SOURCE_DIR / f"{dataset['copy_name']}{newest.suffix.lower()}"
        shutil.copy2(newest, dest)
        return dest

    named = list(SOURCE_DIR.glob(f"{dataset['copy_name']}.*")) + [
        path
        for path in [*WORKSPACE.glob("*"), *SOURCE_DIR.glob("*")]
        if is_usable(path) and match(path.name.casefold())
    ]
    usable = [path for path in named if is_usable(path)]
    if usable:
        return max(usable, key=lambda path: path.stat().st_mtime)
    return None


def read_csv(path: Path) -> list[dict[str, str]]:
    text = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise UnicodeDecodeError("utf-8", b"", 0, 1, f"Could not decode {path}")
    return [{key: cell_text(value) for key, value in row.items()} for row in csv.DictReader(StringIO(text))]


def read_xlsx(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_text(value) or f"column_{index}" for index, value in enumerate(rows[0])]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {}
        for index, header in enumerate(headers):
            record[header] = cell_text(row[index] if index < len(row) else None)
        records.append(record)
    return records


def read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    return read_csv(path)


def pick_field(row: dict[str, str], *names: str) -> str:
    wanted = {name.casefold() for name in names}
    for key, value in row.items():
        if key.strip().casefold() in wanted:
            return (value or "").strip()
    return ""


def pick_contains(row: dict[str, str], *needles: str) -> str:
    for key, value in row.items():
        lowered = key.casefold()
        if all(needle in lowered for needle in needles):
            return (value or "").strip()
    return ""


def empty_payload(dataset: dict, source: str, note: str) -> dict:
    return {
        "meta": {
            "title": dataset["title"],
            "subtitle": dataset["subtitle"],
            "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": source,
            "privacy": "Emails, written answers, and file links were removed before publishing.",
            "note": note,
        },
        "kpis": {
            "completed": 0,
            "latestAt": None,
            "latestLabel": "",
        },
        "byDay": [],
        "people": [],
    }


def build_people(rows: list[dict[str, str]], extras: bool) -> list[dict]:
    best: dict[str, dict] = {}
    for row in rows:
        name = pick_field(row, "name") or pick_contains(row, "what is your name")
        if not name:
            continue
        completed_at = parse_datetime(pick_field(row, "completion time"))
        record = {
            "name": name,
            "completed": True,
            "completedAt": isoformat(completed_at),
            "completedLabel": display_time(completed_at),
        }
        if extras:
            program = pick_contains(row, "guild program")
            status = pick_contains(row, "participation status")
            if program:
                record["program"] = program
            if status:
                record["status"] = status
        key = name.casefold()
        previous = best.get(key)
        if previous is None:
            best[key] = record
            continue
        prev_iso = previous.get("completedAt")
        prev_dt = (
            datetime.strptime(prev_iso, "%Y-%m-%dT%H:%M:%S") if prev_iso else datetime.min
        )
        if completed_at >= prev_dt:
            best[key] = record

    return sorted(
        best.values(),
        key=lambda item: (item["completedAt"] or "", item["name"].casefold()),
        reverse=True,
    )


def summarize(people: list[dict]) -> tuple[dict[str, int], datetime | None]:
    day_counts: dict[str, int] = {}
    latest: datetime | None = None
    for person in people:
        if not person["completedAt"]:
            continue
        day = person["completedAt"][:10]
        day_counts[day] = day_counts.get(day, 0) + 1
        stamp = datetime.strptime(person["completedAt"], "%Y-%m-%dT%H:%M:%S")
        if latest is None or stamp > latest:
            latest = stamp
    return day_counts, latest


def write_dataset(dataset: dict) -> None:
    source = find_source(dataset)
    output = dataset["output"]
    output.parent.mkdir(parents=True, exist_ok=True)
    if source is None:
        payload = empty_payload(
            dataset,
            f"No {dataset['title']} export found",
            "Drop the Forms Excel/CSV into source/ or Downloads, then rerun this script.",
        )
        output.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        print(f"No export found for {dataset['title']}. Wrote empty board to {output}")
        return

    people = build_people(read_rows(source), dataset["extras"])
    day_counts, latest = summarize(people)
    payload = {
        "meta": {
            "title": dataset["title"],
            "subtitle": dataset["subtitle"],
            "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": dataset["source_label"],
            "privacy": "Emails, written answers, and file links were removed before publishing.",
            "note": "Anyone in the Forms export is marked as submitted. Duplicate names keep the latest response.",
        },
        "kpis": {
            "completed": len(people),
            "latestAt": isoformat(latest) if latest else None,
            "latestLabel": display_time(latest) if latest else "",
        },
        "byDay": [{"day": day, "count": day_counts[day]} for day in sorted(day_counts)],
        "people": people,
    }
    output.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {output}")
    print(f"{len(people)} submitted from {source.name}")
    for person in people:
        extra = f" — {person['status']}" if person.get("status") else ""
        print(f"  {person['name']} — {person['completedLabel']}{extra}")


def main() -> None:
    for dataset in DATASETS:
        write_dataset(dataset)


if __name__ == "__main__":
    main()
